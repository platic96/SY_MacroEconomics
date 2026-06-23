"""
포트폴리오 관리 매매일지 - 구글시트용 통합 템플릿
- 8개 시트 구조
- GOOGLEFINANCE 함수 포함 (구글시트 업로드 후 자동 작동)
- 샘플 데이터 포함
- 차트 + 조건부 서식 + 데이터 유효성 검사
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.chart import BarChart, PieChart, LineChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

wb = Workbook()

# ============================================================
# 공통 스타일 정의
# ============================================================
COLOR = {
    'header_bg': '1976D2',      # 헤더 파랑
    'header_fg': 'FFFFFF',
    'sub_header': '90CAF9',     # 서브 헤더
    'profit': 'E8F5E9',         # 수익 연초록
    'loss': 'FFEBEE',           # 손실 연빨강
    'profit_text': '2E7D32',
    'loss_text': 'C62828',
    'highlight': 'FFF9C4',      # 하이라이트 노랑
    'input_bg': 'F5F5F5',       # 입력 셀 회색
    'calc_bg': 'E3F2FD',        # 계산 셀 연파랑
    'border': 'BDBDBD',
    'krw': 'FFF3E0',            # 원화 주황
    'usd': 'E1F5FE',            # 달러 하늘
    'crypto': 'F3E5F5',         # 코인 보라
}

FONT_HEADER = Font(name='맑은 고딕', size=11, bold=True, color=COLOR['header_fg'])
FONT_SUB = Font(name='맑은 고딕', size=10, bold=True, color='212121')
FONT_NORMAL = Font(name='맑은 고딕', size=10)
FONT_BOLD = Font(name='맑은 고딕', size=10, bold=True)
FONT_BIG = Font(name='맑은 고딕', size=14, bold=True)
FONT_HUGE = Font(name='맑은 고딕', size=20, bold=True, color='1976D2')

FILL_HEADER = PatternFill(start_color=COLOR['header_bg'], end_color=COLOR['header_bg'], fill_type='solid')
FILL_SUB = PatternFill(start_color=COLOR['sub_header'], end_color=COLOR['sub_header'], fill_type='solid')
FILL_INPUT = PatternFill(start_color=COLOR['input_bg'], end_color=COLOR['input_bg'], fill_type='solid')
FILL_CALC = PatternFill(start_color=COLOR['calc_bg'], end_color=COLOR['calc_bg'], fill_type='solid')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

thin = Side(border_style='thin', color=COLOR['border'])
medium = Side(border_style='medium', color='1976D2')
BORDER_THIN = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_HEADER = Border(left=thin, right=thin, top=medium, bottom=medium)


def style_header(ws, row, cols, height=30):
    """헤더 행 스타일링"""
    ws.row_dimensions[row].height = height
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_HEADER


def set_col_widths(ws, widths):
    """컬럼 너비 일괄 설정"""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_border_range(ws, cell_range):
    """범위에 테두리 적용"""
    for row in ws[cell_range]:
        for cell in row:
            cell.border = BORDER_THIN


# ============================================================
# 시트 1: 📝 매매입력 (Trade Log)
# ============================================================
ws1 = wb.active
ws1.title = "1.매매입력"
ws1.sheet_properties.tabColor = "1976D2"

headers_1 = [
    "거래일", "자산구분", "시장", "종목명", "티커", "구분",
    "수량", "단가(현지)", "통화", "환율", "수수료(원)", "세금(원)",
    "총액(원화)", "메모"
]
ws1.append(headers_1)
style_header(ws1, 1, len(headers_1))

# 샘플 데이터
sample_trades = [
    ["2026-01-15", "주식", "한국", "삼성전자", "005930", "매수", 100, 78000, "KRW", 1, 1500, 0, None, "1분기 실적 기대"],
    ["2026-02-03", "주식", "미국", "NVIDIA", "NVDA", "매수", 10, 850, "USD", 1380, 2000, 0, None, "AI 트렌드 지속"],
    ["2026-03-20", "ETF", "미국", "Invesco QQQ", "QQQ", "매수", 5, 480, "USD", 1390, 1500, 0, None, "분산 투자"],
    ["2026-04-10", "코인", "암호화폐", "비트코인", "BTC", "매수", 0.05, 95000, "USD", 1400, 5000, 0, None, "장기 보유"],
    ["2026-04-29", "주식", "한국", "삼성전자", "005930", "매도", 50, 92000, "KRW", 1, 1500, 0, None, "일부 차익실현"],
    ["2026-05-15", "주식", "미국", "Apple", "AAPL", "매수", 15, 230, "USD", 1395, 2500, 0, None, "WWDC 기대"],
]

for trade in sample_trades:
    ws1.append(trade)

# 총액(원화) 자동 계산 수식 - 13번째 컬럼 (M)
for row in range(2, len(sample_trades) + 2):
    # 매수: -금액 (지출), 매도: +금액 (수입)
    formula = f'=IF(F{row}="매수", -1, 1) * (G{row} * H{row} * J{row}) - K{row} - L{row}'
    ws1.cell(row=row, column=13).value = formula
    ws1.cell(row=row, column=13).number_format = '#,##0'

# 컬럼 너비
set_col_widths(ws1, [12, 10, 10, 18, 10, 8, 10, 12, 8, 10, 11, 10, 16, 30])

# 데이터 유효성 검사 (드롭다운)
dv_asset = DataValidation(type="list", formula1='"주식,ETF,코인,채권,펀드,기타"', allow_blank=True)
dv_market = DataValidation(type="list", formula1='"한국,미국,중국,일본,유럽,암호화폐,기타"', allow_blank=True)
dv_type = DataValidation(type="list", formula1='"매수,매도"', allow_blank=True)
dv_currency = DataValidation(type="list", formula1='"KRW,USD,JPY,EUR,CNY"', allow_blank=True)

ws1.add_data_validation(dv_asset)
ws1.add_data_validation(dv_market)
ws1.add_data_validation(dv_type)
ws1.add_data_validation(dv_currency)

dv_asset.add('B2:B1000')
dv_market.add('C2:C1000')
dv_type.add('F2:F1000')
dv_currency.add('I2:I1000')

# 조건부 서식: 매수=파랑, 매도=빨강
ws1.conditional_formatting.add('F2:F1000',
    FormulaRule(formula=['$F2="매수"'], fill=PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')))
ws1.conditional_formatting.add('F2:F1000',
    FormulaRule(formula=['$F2="매도"'], fill=PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')))

# 데이터 행 스타일
for row in range(2, len(sample_trades) + 2):
    for col in range(1, 15):
        cell = ws1.cell(row=row, column=col)
        cell.font = FONT_NORMAL
        cell.border = BORDER_THIN
        if col in [7, 8, 10, 11, 12]:
            cell.number_format = '#,##0.00'
            cell.alignment = ALIGN_RIGHT
        elif col == 1:
            cell.number_format = 'yyyy-mm-dd'
            cell.alignment = ALIGN_CENTER
        elif col in [2, 3, 6, 9]:
            cell.alignment = ALIGN_CENTER
        else:
            cell.alignment = ALIGN_LEFT

# 헤더 메모(주석)
ws1.cell(row=1, column=13).comment = Comment("자동계산: 매수=음수, 매도=양수\n수수료/세금 차감 반영", "시스템")
ws1.cell(row=1, column=10).comment = Comment("KRW는 1, USD는 실시간 환율 자동입력 권장\n구글시트: =GOOGLEFINANCE(\"CURRENCY:USDKRW\")", "시스템")

ws1.freeze_panes = "A2"


# ============================================================
# 시트 2: 📊 보유현황 (Holdings)
# ============================================================
ws2 = wb.create_sheet("2.보유현황")
ws2.sheet_properties.tabColor = "4CAF50"

# 안내문
ws2['A1'] = "📊 실시간 보유 종목 현황"
ws2['A1'].font = FONT_HUGE
ws2.merge_cells('A1:K1')
ws2.row_dimensions[1].height = 35

ws2['A2'] = "※ 종목을 추가하려면 아래 표에 티커를 입력하세요. 보유수량·평균단가는 매매입력 시트에서 자동 집계됩니다."
ws2['A2'].font = Font(name='맑은 고딕', size=9, italic=True, color='616161')
ws2.merge_cells('A2:K2')

headers_2 = [
    "종목명", "티커", "자산구분", "시장", "통화",
    "보유수량", "평균단가", "현재가", "평가금액(원)", "평가손익(원)", "수익률(%)"
]
ws2.append([])  # 빈 행
for col, h in enumerate(headers_2, 1):
    ws2.cell(row=4, column=col, value=h)
style_header(ws2, 4, len(headers_2))

# 샘플 보유종목 (매매입력에서 자동집계되도록 수식 작성)
holdings = [
    ["삼성전자", "005930", "주식", "한국", "KRW"],
    ["NVIDIA", "NVDA", "주식", "미국", "USD"],
    ["Invesco QQQ", "QQQ", "ETF", "미국", "USD"],
    ["비트코인", "BTC", "코인", "암호화폐", "USD"],
    ["Apple", "AAPL", "주식", "미국", "USD"],
]

for i, h in enumerate(holdings, 5):
    ws2.cell(row=i, column=1, value=h[0])
    ws2.cell(row=i, column=2, value=h[1])
    ws2.cell(row=i, column=3, value=h[2])
    ws2.cell(row=i, column=4, value=h[3])
    ws2.cell(row=i, column=5, value=h[4])

    # 보유수량 = 매수합계 - 매도합계
    ws2.cell(row=i, column=6,
        value=f'=SUMIFS(\'1.매매입력\'!G:G,\'1.매매입력\'!E:E,B{i},\'1.매매입력\'!F:F,"매수") - SUMIFS(\'1.매매입력\'!G:G,\'1.매매입력\'!E:E,B{i},\'1.매매입력\'!F:F,"매도")')

    # 평균단가 = (매수금액 / 매수수량) - 가중평균
    ws2.cell(row=i, column=7,
        value=f'=IFERROR(SUMPRODUCT((\'1.매매입력\'!E$2:E$1000=B{i})*(\'1.매매입력\'!F$2:F$1000="매수")*\'1.매매입력\'!G$2:G$1000*\'1.매매입력\'!H$2:H$1000) / SUMIFS(\'1.매매입력\'!G:G,\'1.매매입력\'!E:E,B{i},\'1.매매입력\'!F:F,"매수"), 0)')

    # 현재가 - GOOGLEFINANCE (구글시트 업로드 후 작동)
    if h[3] == "한국":
        ws2.cell(row=i, column=8, value=f'=GOOGLEFINANCE("KRX:"&B{i})')
    elif h[3] == "미국":
        ws2.cell(row=i, column=8, value=f'=GOOGLEFINANCE(B{i})')
    elif h[3] == "암호화폐":
        ws2.cell(row=i, column=8, value=f'=GOOGLEFINANCE("CURRENCY:"&B{i}&"USD")')
    else:
        ws2.cell(row=i, column=8, value=0)

    # 평가금액(원) = 보유수량 × 현재가 × 환율
    ws2.cell(row=i, column=9,
        value=f'=IFERROR(F{i}*H{i}*VLOOKUP(E{i},\'8.설정\'!$B$5:$C$10,2,FALSE), 0)')

    # 평가손익 = 평가금액 - 매수원가
    ws2.cell(row=i, column=10,
        value=f'=I{i} - F{i}*G{i}*VLOOKUP(E{i},\'8.설정\'!$B$5:$C$10,2,FALSE)')

    # 수익률
    ws2.cell(row=i, column=11,
        value=f'=IFERROR(J{i}/(F{i}*G{i}*VLOOKUP(E{i},\'8.설정\'!$B$5:$C$10,2,FALSE)), 0)')

# 합계 행
total_row = 5 + len(holdings)
ws2.cell(row=total_row, column=1, value="합계").font = FONT_BOLD
ws2.cell(row=total_row, column=9, value=f'=SUM(I5:I{total_row-1})')
ws2.cell(row=total_row, column=10, value=f'=SUM(J5:J{total_row-1})')
ws2.cell(row=total_row, column=11, value=f'=IFERROR(J{total_row}/(I{total_row}-J{total_row}), 0)')

for col in range(1, 12):
    ws2.cell(row=total_row, column=col).fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    ws2.cell(row=total_row, column=col).font = FONT_BOLD
    ws2.cell(row=total_row, column=col).border = Border(top=medium, bottom=medium, left=thin, right=thin)

# 셀 서식
for row in range(5, total_row + 1):
    for col in range(1, 12):
        cell = ws2.cell(row=row, column=col)
        cell.border = BORDER_THIN
        if col == 6:
            cell.number_format = '#,##0.0000'
            cell.alignment = ALIGN_RIGHT
        elif col in [7, 8]:
            cell.number_format = '#,##0.00'
            cell.alignment = ALIGN_RIGHT
        elif col in [9, 10]:
            cell.number_format = '#,##0'
            cell.alignment = ALIGN_RIGHT
        elif col == 11:
            cell.number_format = '0.00%'
            cell.alignment = ALIGN_RIGHT
        elif col in [3, 4, 5]:
            cell.alignment = ALIGN_CENTER

# 조건부 서식: 수익/손실 색상
ws2.conditional_formatting.add(f'J5:J{total_row}',
    CellIsRule(operator='greaterThan', formula=['0'], fill=PatternFill(start_color=COLOR['profit'], end_color=COLOR['profit'], fill_type='solid'), font=Font(color=COLOR['profit_text'], bold=True)))
ws2.conditional_formatting.add(f'J5:J{total_row}',
    CellIsRule(operator='lessThan', formula=['0'], fill=PatternFill(start_color=COLOR['loss'], end_color=COLOR['loss'], fill_type='solid'), font=Font(color=COLOR['loss_text'], bold=True)))
ws2.conditional_formatting.add(f'K5:K{total_row}',
    CellIsRule(operator='greaterThan', formula=['0'], font=Font(color=COLOR['profit_text'], bold=True)))
ws2.conditional_formatting.add(f'K5:K{total_row}',
    CellIsRule(operator='lessThan', formula=['0'], font=Font(color=COLOR['loss_text'], bold=True)))

set_col_widths(ws2, [18, 10, 10, 10, 8, 12, 12, 12, 16, 14, 10])
ws2.freeze_panes = "A5"


# ============================================================
# 시트 3: 💵 실현손익 (Realized P&L)
# ============================================================
ws3 = wb.create_sheet("3.실현손익")
ws3.sheet_properties.tabColor = "FF9800"

ws3['A1'] = "💵 실현 손익 (양도세 계산용)"
ws3['A1'].font = FONT_HUGE
ws3.merge_cells('A1:J1')
ws3.row_dimensions[1].height = 35

headers_3 = [
    "매도일", "종목명", "티커", "매도수량", "매도단가",
    "평균매수단가", "보유기간(일)", "통화", "실현손익(원)", "양도세 예상"
]
ws3.append([])
for col, h in enumerate(headers_3, 1):
    ws3.cell(row=3, column=col, value=h)
style_header(ws3, 3, len(headers_3))

# 매도 거래만 가져와서 실현손익 계산
realized_samples = [
    ["2026-04-29", "삼성전자", "005930", 50, 92000, 78000, 104, "KRW"],
]

for i, r in enumerate(realized_samples, 4):
    for col, val in enumerate(r, 1):
        ws3.cell(row=i, column=col, value=val)
    # 실현손익(원)
    ws3.cell(row=i, column=9,
        value=f'=(E{i}-F{i})*D{i}*VLOOKUP(H{i},\'8.설정\'!$B$5:$C$10,2,FALSE)')
    # 양도세 예상 (해외주식 22%, 국내 0%, 코인 미정)
    ws3.cell(row=i, column=10,
        value=f'=IF(I{i}>0, IF(VLOOKUP(B{i},\'2.보유현황\'!A:D,4,FALSE)="미국", MAX(0,(I{i}-2500000))*0.22, 0), 0)')

for row in range(4, 4 + len(realized_samples) + 5):
    for col in range(1, 11):
        cell = ws3.cell(row=row, column=col)
        cell.border = BORDER_THIN
        cell.font = FONT_NORMAL
        if col == 1:
            cell.number_format = 'yyyy-mm-dd'
            cell.alignment = ALIGN_CENTER
        elif col in [4, 5, 6, 7]:
            cell.number_format = '#,##0.00'
            cell.alignment = ALIGN_RIGHT
        elif col in [9, 10]:
            cell.number_format = '#,##0'
            cell.alignment = ALIGN_RIGHT
        elif col in [3, 8]:
            cell.alignment = ALIGN_CENTER

ws3.conditional_formatting.add('I4:I100',
    CellIsRule(operator='greaterThan', formula=['0'], fill=PatternFill(start_color=COLOR['profit'], end_color=COLOR['profit'], fill_type='solid'), font=Font(color=COLOR['profit_text'], bold=True)))
ws3.conditional_formatting.add('I4:I100',
    CellIsRule(operator='lessThan', formula=['0'], fill=PatternFill(start_color=COLOR['loss'], end_color=COLOR['loss'], fill_type='solid'), font=Font(color=COLOR['loss_text'], bold=True)))

# 합계
sum_row = 4 + len(realized_samples) + 1
ws3.cell(row=sum_row, column=1, value="연간 합계").font = FONT_BOLD
ws3.cell(row=sum_row, column=9, value=f'=SUM(I4:I{sum_row-1})')
ws3.cell(row=sum_row, column=10, value=f'=SUM(J4:J{sum_row-1})')
for col in range(1, 11):
    ws3.cell(row=sum_row, column=col).fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    ws3.cell(row=sum_row, column=col).font = FONT_BOLD

set_col_widths(ws3, [12, 18, 10, 10, 12, 14, 12, 8, 16, 14])
ws3.freeze_panes = "A4"


# ============================================================
# 시트 4: 💰 배당내역 (Dividends)
# ============================================================
ws4 = wb.create_sheet("4.배당내역")
ws4.sheet_properties.tabColor = "9C27B0"

ws4['A1'] = "💰 배당 수령 내역"
ws4['A1'].font = FONT_HUGE
ws4.merge_cells('A1:I1')
ws4.row_dimensions[1].height = 35

headers_4 = [
    "지급일", "종목명", "티커", "보유수량", "주당배당",
    "통화", "환율", "세전배당(원)", "세후배당(원)"
]
ws4.append([])
for col, h in enumerate(headers_4, 1):
    ws4.cell(row=3, column=col, value=h)
style_header(ws4, 3, len(headers_4))

dividends = [
    ["2026-03-15", "삼성전자", "005930", 100, 360, "KRW", 1, None, None],
    ["2026-03-31", "Apple", "AAPL", 15, 0.25, "USD", 1395, None, None],
    ["2026-04-30", "Invesco QQQ", "QQQ", 5, 0.65, "USD", 1400, None, None],
]

for i, d in enumerate(dividends, 4):
    for col, val in enumerate(d, 1):
        if val is not None:
            ws4.cell(row=i, column=col, value=val)
    ws4.cell(row=i, column=8, value=f'=D{i}*E{i}*G{i}')
    ws4.cell(row=i, column=9, value=f'=H{i}*IF(F{i}="KRW", 0.846, 0.85)')

for row in range(4, 4 + len(dividends)):
    for col in range(1, 10):
        cell = ws4.cell(row=row, column=col)
        cell.border = BORDER_THIN
        cell.font = FONT_NORMAL
        if col == 1:
            cell.number_format = 'yyyy-mm-dd'
            cell.alignment = ALIGN_CENTER
        elif col in [4, 5, 7]:
            cell.number_format = '#,##0.0000'
            cell.alignment = ALIGN_RIGHT
        elif col in [8, 9]:
            cell.number_format = '#,##0'
            cell.alignment = ALIGN_RIGHT
        elif col in [3, 6]:
            cell.alignment = ALIGN_CENTER

# 합계 행
sum_row = 4 + len(dividends) + 1
ws4.cell(row=sum_row, column=1, value="누적 합계").font = FONT_BOLD
ws4.cell(row=sum_row, column=8, value=f'=SUM(H4:H{sum_row-1})')
ws4.cell(row=sum_row, column=9, value=f'=SUM(I4:I{sum_row-1})')
for col in range(1, 10):
    ws4.cell(row=sum_row, column=col).fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    ws4.cell(row=sum_row, column=col).font = FONT_BOLD

ws4.cell(row=3, column=9).comment = Comment("한국 배당세 15.4%\n미국 배당세 15%", "시스템")

set_col_widths(ws4, [12, 18, 10, 10, 12, 8, 10, 16, 16])
ws4.freeze_panes = "A4"


# ============================================================
# 시트 5: 📓 매매일지 (Trading Journal)
# ============================================================
ws5 = wb.create_sheet("5.매매일지")
ws5.sheet_properties.tabColor = "F44336"

ws5['A1'] = "📓 매매일지 - 왜 샀고 왜 팔았는가"
ws5['A1'].font = FONT_HUGE
ws5.merge_cells('A1:J1')
ws5.row_dimensions[1].height = 35

ws5['A2'] = "※ 거래 직전·직후에 작성하면 감정적 매매가 줄고 실력이 빠르게 늡니다."
ws5['A2'].font = Font(name='맑은 고딕', size=9, italic=True, color='616161')
ws5.merge_cells('A2:J2')

headers_5 = [
    "일자", "종목", "티커", "구분", "진입가", "손절가", "목표가",
    "진입 근거", "결과 회고 (잘한점/못한점)", "교훈"
]
ws5.append([])
for col, h in enumerate(headers_5, 1):
    ws5.cell(row=4, column=col, value=h)
style_header(ws5, 4, len(headers_5))

journal_samples = [
    ["2026-02-03", "NVIDIA", "NVDA", "진입", 850, 780, 1100,
     "AI 데이터센터 CAPEX 폭증, FY26 가이던스 상향, Blackwell 양산 시작",
     "진입 타이밍 좋았음. 다만 분할매수했어야 했음.",
     "단번에 풀매수 X, 3차에 나눠 진입"],
    ["2026-04-29", "삼성전자", "005930", "청산(일부)", 78000, 72000, 95000,
     "HBM4 완판 + 메모리 이익률 74%, 1분기 실적 서프라이즈",
     "+18% 차익실현 성공. 그러나 너무 일찍 절반 매도. 더 들고 갔어야.",
     "추세 살아있을 때는 트레일링스탑 사용"],
]

for i, j in enumerate(journal_samples, 5):
    for col, val in enumerate(j, 1):
        ws5.cell(row=i, column=col, value=val)

for row in range(5, 5 + len(journal_samples) + 5):
    ws5.row_dimensions[row].height = 60
    for col in range(1, 11):
        cell = ws5.cell(row=row, column=col)
        cell.border = BORDER_THIN
        cell.font = FONT_NORMAL
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        if col == 1:
            cell.number_format = 'yyyy-mm-dd'
            cell.alignment = Alignment(horizontal='center', vertical='top')
        elif col in [3, 4]:
            cell.alignment = Alignment(horizontal='center', vertical='top')
        elif col in [5, 6, 7]:
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right', vertical='top')

dv_journal_type = DataValidation(type="list", formula1='"진입,청산,청산(일부),관망,추가매수"', allow_blank=True)
ws5.add_data_validation(dv_journal_type)
dv_journal_type.add('D5:D1000')

set_col_widths(ws5, [12, 14, 10, 12, 10, 10, 10, 32, 32, 25])
ws5.freeze_panes = "A5"


# ============================================================
# 시트 6: 📈 대시보드 (Dashboard)
# ============================================================
ws6 = wb.create_sheet("6.대시보드")
ws6.sheet_properties.tabColor = "00BCD4"

ws6['A1'] = "📈 포트폴리오 종합 대시보드"
ws6['A1'].font = FONT_HUGE
ws6.merge_cells('A1:H1')
ws6.row_dimensions[1].height = 40

# KPI 카드 영역
kpi_data = [
    ("총 평가금액", f"=\'2.보유현황\'!I{5 + len(holdings)}", "#,##0", '1976D2'),
    ("총 평가손익", f"=\'2.보유현황\'!J{5 + len(holdings)}", "#,##0", '4CAF50'),
    ("총 수익률", f"=\'2.보유현황\'!K{5 + len(holdings)}", "0.00%", 'FF9800'),
    ("실현손익(YTD)", f"=\'3.실현손익\'!I{4 + len(realized_samples) + 1}", "#,##0", '9C27B0'),
    ("배당수령(누적)", f"=\'4.배당내역\'!I{4 + len(dividends) + 1}", "#,##0", 'F44336'),
    ("종목 수", f"=COUNTA(\'2.보유현황\'!B5:B{4 + len(holdings)})", "0", '00BCD4'),
]

# 2행 3열 KPI 카드
ws6['A3'] = "🏆 핵심 지표"
ws6['A3'].font = FONT_BIG
ws6.merge_cells('A3:H3')

for idx, (label, formula, fmt, color) in enumerate(kpi_data):
    col_start = (idx % 3) * 3 + 1  # 1, 4, 7
    row_start = 5 + (idx // 3) * 4

    # 라벨
    ws6.cell(row=row_start, column=col_start, value=label)
    ws6.cell(row=row_start, column=col_start).font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    ws6.cell(row=row_start, column=col_start).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    ws6.cell(row=row_start, column=col_start).alignment = ALIGN_CENTER
    ws6.merge_cells(start_row=row_start, start_column=col_start, end_row=row_start, end_column=col_start+1)

    # 값
    ws6.cell(row=row_start+1, column=col_start, value=formula)
    ws6.cell(row=row_start+1, column=col_start).font = Font(name='맑은 고딕', size=16, bold=True)
    ws6.cell(row=row_start+1, column=col_start).fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    ws6.cell(row=row_start+1, column=col_start).number_format = fmt
    ws6.cell(row=row_start+1, column=col_start).alignment = ALIGN_CENTER
    ws6.merge_cells(start_row=row_start+1, start_column=col_start, end_row=row_start+2, end_column=col_start+1)

    ws6.row_dimensions[row_start].height = 25
    ws6.row_dimensions[row_start+1].height = 28

# 차트용 데이터 영역 (보유현황 참조)
ws6['A14'] = "📊 종목별 비중"
ws6['A14'].font = FONT_BIG

# 차트 데이터를 명시적으로 준비 (PieChart는 직접 참조가 어려움)
for i, h in enumerate(holdings):
    ws6.cell(row=15+i, column=1, value=h[0])
    ws6.cell(row=15+i, column=2, value=f"=\'2.보유현황\'!I{5+i}")
    ws6.cell(row=15+i, column=2).number_format = '#,##0'

# 도넛 차트
pie = DoughnutChart()
pie.title = "종목별 평가금액 비중"
labels = Reference(ws6, min_col=1, min_row=15, max_row=14+len(holdings))
data = Reference(ws6, min_col=2, min_row=14, max_row=14+len(holdings))
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.height = 9
pie.width = 12
pie.dataLabels = DataLabelList(showPercent=True)
ws6.add_chart(pie, "D14")

# 자산구분 비중
ws6['A22'] = "🥧 자산 구분별 비중"
ws6['A22'].font = FONT_BIG

asset_types = ["주식", "ETF", "코인", "채권", "펀드"]
for i, at in enumerate(asset_types):
    ws6.cell(row=23+i, column=1, value=at)
    ws6.cell(row=23+i, column=2,
        value=f'=SUMIFS(\'2.보유현황\'!I:I, \'2.보유현황\'!C:C, "{at}")')
    ws6.cell(row=23+i, column=2).number_format = '#,##0'

pie2 = PieChart()
pie2.title = "자산구분별 비중"
labels2 = Reference(ws6, min_col=1, min_row=23, max_row=22+len(asset_types))
data2 = Reference(ws6, min_col=2, min_row=22, max_row=22+len(asset_types))
pie2.add_data(data2, titles_from_data=True)
pie2.set_categories(labels2)
pie2.height = 9
pie2.width = 12
pie2.dataLabels = DataLabelList(showPercent=True)
ws6.add_chart(pie2, "D22")

set_col_widths(ws6, [14, 14, 2, 14, 14, 2, 14, 14])

# 헤더 행 인쇄 영역
ws6.freeze_panes = "A4"


# ============================================================
# 시트 7: 📅 월별 수익률 (Performance)
# ============================================================
ws7 = wb.create_sheet("7.월별수익률")
ws7.sheet_properties.tabColor = "FFC107"

ws7['A1'] = "📅 월별/연도별 성과 추적"
ws7['A1'].font = FONT_HUGE
ws7.merge_cells('A1:H1')
ws7.row_dimensions[1].height = 35

headers_7 = ["월", "시작자산(원)", "입금", "출금", "종료자산(원)", "월수익(원)", "월수익률(%)", "YTD수익률(%)"]
ws7.append([])
for col, h in enumerate(headers_7, 1):
    ws7.cell(row=3, column=col, value=h)
style_header(ws7, 3, len(headers_7))

months_2026 = [
    "2026-01", "2026-02", "2026-03", "2026-04", "2026-05",
    "2026-06", "2026-07", "2026-08", "2026-09", "2026-10", "2026-11", "2026-12"
]

# 샘플: 1월부터 4월까지만 입력
sample_returns = [
    [100000000, 5000000, 0, 108000000],
    [108000000, 0, 0, 112000000],
    [112000000, 3000000, 0, 118000000],
    [118000000, 0, 2000000, 125000000],
]

for i, m in enumerate(months_2026, 4):
    ws7.cell(row=i, column=1, value=m)
    if i - 4 < len(sample_returns):
        ws7.cell(row=i, column=2, value=sample_returns[i-4][0])
        ws7.cell(row=i, column=3, value=sample_returns[i-4][1])
        ws7.cell(row=i, column=4, value=sample_returns[i-4][2])
        ws7.cell(row=i, column=5, value=sample_returns[i-4][3])

    # 월수익
    ws7.cell(row=i, column=6, value=f'=IFERROR(E{i}-B{i}-C{i}+D{i},"")')
    # 월수익률
    ws7.cell(row=i, column=7, value=f'=IFERROR(F{i}/(B{i}+C{i}),"")')
    # YTD수익률
    if i == 4:
        ws7.cell(row=i, column=8, value=f'=IFERROR(G{i},"")')
    else:
        ws7.cell(row=i, column=8, value=f'=IFERROR((1+H{i-1})*(1+G{i})-1,"")')

for row in range(4, 16):
    for col in range(1, 9):
        cell = ws7.cell(row=row, column=col)
        cell.border = BORDER_THIN
        cell.font = FONT_NORMAL
        if col == 1:
            cell.alignment = ALIGN_CENTER
            cell.font = FONT_BOLD
        elif col in [2, 3, 4, 5, 6]:
            cell.number_format = '#,##0'
            cell.alignment = ALIGN_RIGHT
        elif col in [7, 8]:
            cell.number_format = '0.00%'
            cell.alignment = ALIGN_RIGHT

ws7.conditional_formatting.add('G4:H15',
    CellIsRule(operator='greaterThan', formula=['0'], font=Font(color=COLOR['profit_text'], bold=True)))
ws7.conditional_formatting.add('G4:H15',
    CellIsRule(operator='lessThan', formula=['0'], font=Font(color=COLOR['loss_text'], bold=True)))

# 월별 수익률 차트
ws7['J3'] = "📊 월별 수익률 추이"
ws7['J3'].font = FONT_BIG

line = LineChart()
line.title = "월별 수익률 (%)"
line.y_axis.title = "수익률"
line.x_axis.title = "월"
data_line = Reference(ws7, min_col=7, min_row=3, max_row=15)
cats_line = Reference(ws7, min_col=1, min_row=4, max_row=15)
line.add_data(data_line, titles_from_data=True)
line.set_categories(cats_line)
line.height = 9
line.width = 16
ws7.add_chart(line, "J5")

set_col_widths(ws7, [10, 16, 12, 12, 16, 14, 12, 14, 2, 18])
ws7.freeze_panes = "A4"


# ============================================================
# 시트 8: ⚙️ 설정 (Settings)
# ============================================================
ws8 = wb.create_sheet("8.설정")
ws8.sheet_properties.tabColor = "607D8B"

ws8['A1'] = "⚙️ 설정 - 환율 / 종목마스터 / 세율"
ws8['A1'].font = FONT_HUGE
ws8.merge_cells('A1:F1')
ws8.row_dimensions[1].height = 35

# 환율
ws8['A3'] = "💱 환율 (구글시트 업로드 시 자동 갱신)"
ws8['A3'].font = FONT_BIG
ws8.merge_cells('A3:F3')

ws8['B4'] = "통화"
ws8['C4'] = "환율(원)"
ws8['D4'] = "비고"
for col, val in enumerate([ws8['B4'].value, ws8['C4'].value, ws8['D4'].value], 2):
    cell = ws8.cell(row=4, column=col)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER_THIN

currencies = [
    ("KRW", 1, "기준 통화"),
    ("USD", '=GOOGLEFINANCE("CURRENCY:USDKRW")', "미국 달러"),
    ("JPY", '=GOOGLEFINANCE("CURRENCY:JPYKRW")', "일본 엔"),
    ("EUR", '=GOOGLEFINANCE("CURRENCY:EURKRW")', "유로"),
    ("CNY", '=GOOGLEFINANCE("CURRENCY:CNYKRW")', "위안"),
    ("HKD", '=GOOGLEFINANCE("CURRENCY:HKDKRW")', "홍콩 달러"),
]

for i, c in enumerate(currencies, 5):
    ws8.cell(row=i, column=2, value=c[0]).alignment = ALIGN_CENTER
    ws8.cell(row=i, column=3, value=c[1]).number_format = '#,##0.00'
    ws8.cell(row=i, column=3).alignment = ALIGN_RIGHT
    ws8.cell(row=i, column=4, value=c[2]).alignment = ALIGN_LEFT
    for col in range(2, 5):
        ws8.cell(row=i, column=col).border = BORDER_THIN
        ws8.cell(row=i, column=col).font = FONT_NORMAL

# 임시 환율값 (구글시트 업로드 전 사용)
ws8.cell(row=6, column=3, value=1400).number_format = '#,##0.00'  # USD 임시값

# 세율 설정
ws8['A13'] = "💸 세율 설정"
ws8['A13'].font = FONT_BIG
ws8.merge_cells('A13:F13')

tax_rates = [
    ("구분", "세율", "비고"),
    ("국내주식 양도세", "0%", "대주주 제외, 일반투자자 비과세"),
    ("해외주식 양도세", "22%", "연 250만원 공제 후 22% (지방세 포함)"),
    ("국내 배당세", "15.4%", "원천징수 (소득세 14% + 지방세 1.4%)"),
    ("해외 배당세", "15%", "미국 기준 (현지 원천징수)"),
    ("코인 양도세", "22%", "2027년부터 시행 예정 (250만원 공제 후)"),
]

for i, t in enumerate(tax_rates, 14):
    for col, val in enumerate(t, 2):
        cell = ws8.cell(row=i, column=col, value=val)
        cell.border = BORDER_THIN
        if i == 14:
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
        else:
            cell.font = FONT_NORMAL
            if col == 2:
                cell.font = FONT_BOLD
            cell.alignment = ALIGN_LEFT if col == 4 else ALIGN_CENTER

# 종목 마스터 (필요시 추가)
ws8['A22'] = "📋 종목 마스터 (선택사항 - 종목 정보 관리용)"
ws8['A22'].font = FONT_BIG
ws8.merge_cells('A22:F22')

master_headers = ["티커", "종목명", "시장", "섹터", "메모"]
for col, h in enumerate(master_headers, 2):
    cell = ws8.cell(row=23, column=col, value=h)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER_THIN

master_data = [
    ["005930", "삼성전자", "한국", "반도체", "메모리 1위"],
    ["NVDA", "NVIDIA", "미국", "반도체", "AI GPU 독점"],
    ["AAPL", "Apple", "미국", "테크", "스마트폰/서비스"],
    ["QQQ", "Invesco QQQ", "미국", "ETF", "나스닥100 추종"],
    ["BTC", "비트코인", "암호화폐", "코인", "디지털 금"],
]

for i, m in enumerate(master_data, 24):
    for col, val in enumerate(m, 2):
        cell = ws8.cell(row=i, column=col, value=val)
        cell.border = BORDER_THIN
        cell.font = FONT_NORMAL
        cell.alignment = ALIGN_CENTER if col != 6 else ALIGN_LEFT

set_col_widths(ws8, [2, 18, 16, 16, 14, 28])
ws8.freeze_panes = "A2"


# ============================================================
# 시트 정렬 및 저장
# ============================================================
# 매매입력 시트를 첫 번째로 설정
wb.active = 0

output_path = "/home/user/SY_MacroEconomics/포트폴리오_매매일지.xlsx"
wb.save(output_path)
print(f"✅ 저장 완료: {output_path}")
print(f"   시트 개수: {len(wb.sheetnames)}")
print(f"   시트 목록: {wb.sheetnames}")
