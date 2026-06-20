from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "IREN 주요 이벤트 타임라인"

headers = ["일자", "주가($)", "구분", "주요 내용 및 계약 상세", "시장 반응 및 주가 영향"]

data = [
    ["2월 5일", 39, "실적 발표",
     "FY26 Q2 실적 발표 (어닝 미스)\n- 매출 $184.7M (예상치 $229.6M 하회)\n- 오클라호마에 1.6GW 규모 신규 AI 캠퍼스 부지 확보 발표",
     "매출 미달로 단기 주가 조정을 받았으나, 기가와트(GW)급 전력망 영토를 크게 넓히며 미래 기초체력을 다짐."],

    ["5월 1일", 45, "인프라",
     "텍사스 Sweetwater 1 데이터 센터 전력 가동\n- 1.4GW급 초대형 변전소를 텍사스 전력망(ERCOT)에 성공적으로 연결",
     "본격적인 빅테크 향 대규모 GPU 가동 준비 완료 신호탄."],

    ["5월 5일", 54, "인수 (M&A)",
     "AI 클라우드 전문 기업 'Mirantis(미란티스)' 인수\n- 엔비디아 칩을 제어·운영할 소프트웨어 및 엔지니어링 역량 내재화",
     "하드웨어(데이터 센터)를 넘어 소프트웨어 기술까지 갖춘 완전체 AI 기업으로 진화 개시."],

    ["5월 7일", 56, "초대형 계약",
     "엔비디아(NVIDIA)와 5년간 34억 달러 규모 AI 클라우드 계약 체결\n- 최신 '블랙웰(Blackwell)' GPU 전격 배치\n- 엔비디아에 $2.1bn 규모 주식 매수 권리(워런트) 부여",
     "올해 최고 호재. 비트코인과 무관하게 주가가 수십 달러 선에서 75달러 부근까지 폭등하게 만든 일등 공신."],

    ["5월 중순 (11일~14일)", 55, "자금 조달",
     "$3.0 Billion (약 4조 원) 규모 전환사채(Convertible Notes) 발행 완료\n- 초기 $2.0bn 계획에서 수요 폭발로 $3.0bn까지 증액 발행",
     "엔비디아 칩을 사기 위한 막대한 실탄을 확보했으나, 향후 주식 전환에 따른 '주가 희석 우려'가 시장에 처음 반영됨."],

    ["5월 26일", 59, "매출 로드맵",
     "마이크로소프트(MS) 계약 진척 및 매출 목표 상향\n- MS향 공급 계약 기반으로 선수금 $1.9bn 확보 상태임을 재확인\n- 2026년 말 기준 연간 반복 매출(ARR) 목표 $4.4bn으로 상향 조정",
     "확실한 백랍(MS)이 있음을 증명하며 시장의 신뢰를 공고히 함."],

    ["6월 1일", 65, "금융 계약",
     "$3.65 Billion (약 5조 원) 규모 GPU 전용 금융 조달(Financing) 마감\n- 투자 적격 등급으로 대규모 대출 및 리스 계약 완료",
     "엔비디아 칩 인도 대금을 완납할 수 있게 됨. 대규모 자본 지출(CapEx)의 피크를 찍음."],

    ["6월 3일", 65, "글로벌 확장",
     "호주에 최초의 데이터 센터 캠퍼스 건설 발표\n- 남호주 지역에 800MW 규모로 구축 계약",
     "미국(텍사스, 오클라호마)을 넘어 아시아·태평양(APAC) 지역까지 인프라 영토 확장."],

    ["6월 15일", 60, "글로벌 확장",
     "유럽 'Nostrum Group' 인수 완료\n- 스페인 중심 490MW 전력망 및 기가와트급 개발 파이프라인 확보",
     "미국, 호주에 이어 유럽 AI 클라우드 시장까지 진출 완료하며 글로벌 인프라 대장주 굳히기."],
]

header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

cell_font = Font(name="맑은 고딕", size=10)
cell_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
date_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
price_align = Alignment(horizontal="center", vertical="center")

thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

category_colors = {
    "실적 발표": "FFF2CC",
    "인프라": "E2EFDA",
    "인수 (M&A)": "FCE4D6",
    "초대형 계약": "FFD966",
    "자금 조달": "DDEBF7",
    "매출 로드맵": "E2EFDA",
    "금융 계약": "DDEBF7",
    "글로벌 확장": "EDEDED",
}

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font
        cell.border = border
        if col_idx == 1:
            cell.alignment = date_align
            cell.font = Font(name="맑은 고딕", size=10, bold=True)
        elif col_idx == 2:
            cell.alignment = price_align
            cell.number_format = '"$"#,##0'
            cell.font = Font(name="맑은 고딕", size=10, bold=True, color="C00000")
        elif col_idx == 3:
            cell.alignment = date_align
            cell.font = Font(name="맑은 고딕", size=10, bold=True)
            color = category_colors.get(value, "FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        else:
            cell.alignment = cell_align

column_widths = [22, 10, 16, 60, 50]
for col_idx, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 30
for row_idx in range(2, len(data) + 2):
    ws.row_dimensions[row_idx].height = 80

ws.freeze_panes = "A2"

output_path = "/home/user/SY_MacroEconomics/IREN_타임라인.xlsx"
wb.save(output_path)
print(f"저장 완료: {output_path}")
